"""Detect a PDF's tables and write them to a spreadsheet.

A conversion filter changes a document's CONTAINER, never its model, so nothing
in a page-description-to-spreadsheet conversion ever decides what a cell is.
The cells therefore come from the page's own channels: the glyph-accurate line
segments the label binder clusters, and the painted paths the field detector
walks.

Five things the geometry forces:

1. **A table is a REGION, not a page.** One page can hold prose, a table, more
   prose, and a second table with a different column count. A region is a run of
   consecutive lines over which one column structure holds; two column
   structures are two regions and two sheets.

2. **Columns come from cell edges repeated across rows.** A table drawn with no
   vertical rules has no other source, and that is the commonest shape. A column
   is anchored by whichever edge its cells actually share — left for a text
   column, right for a numeric one — because alignment is a property of the
   column, not of the page.

3. **Vertical rules corroborate and split; they never establish.** A rule
   standing inside a region with no column beside it bounds a column every row
   leaves empty. A rule is not the source of a column, because the unruled table
   has none.

4. **A row is a band with consistent column OCCUPANCY.** Occupancy is a tight
   test against a column's own anchor; placement, once a region is confirmed, is
   nearest-column and total. A segment-count test admits a two-word heading;
   occupancy does not.

5. **A rotated page reasons in un-rotated user space.** That is where the
   content stream draws, and every rect these channels report is in it.

The page's own strokes are classified here rather than through the field
detector's classifier: that vocabulary is form fields, where a tall thin stroke
is a box edge rather than a column boundary.
"""

from __future__ import annotations

import datetime
import re
from pathlib import Path

import pikepdf

from engine import bidi
from engine.form_detect import (
    BORDER_AREA_FRACTION,
    MIN_RULE_WIDTH,
    RULE_THICKNESS_FLOOR,
    _crop_box,
    _page_segments,
    _page_shapes,
)

# Cell edges within this distance are the same column anchor.
COLUMN_TOLERANCE = 6.0
# How many rows must share an anchor before it is a column rather than a
# coincidence.
MIN_COLUMN_ROWS = 3
# Below these a region is indistinguishable from aligned prose.
MIN_REGION_ROWS = 3
MIN_REGION_COLUMNS = 2
# Two real columns never start closer together than this.
MIN_COLUMN_SEPARATION = 14.0
# Baselines within this distance are the same row, floored so that a line of
# mixed sizes still clusters.
ROW_TOLERANCE_FLOOR = 2.0
ROW_TOLERANCE_EM = 0.3
# How far a segment must reach past a column's boundary to be spanning it rather
# than merely running close to it.
MERGE_OVERLAP = 4.0
# A caption sits no further above a region than this many row pitches.
CAPTION_PITCHES = 3.0
# The format's own limit on a sheet name.
SHEET_NAME_LIMIT = 31
# Characters a sheet name may not carry.
_SHEET_NAME_BANNED = re.compile(r"[\[\]:*?/\\]")
SHEET_PER = ("table", "page")

_ISO_DATE = re.compile(r"^(\d{4})-(\d{2})-(\d{2})$")
# A currency decoration is a short run of characters that are neither
# alphanumeric nor part of a number's own punctuation. A three-letter code is
# deliberately NOT one: it cannot be told from a label like "Q1" without
# guessing, and a guess here silently retypes a string as a number.
_SYMBOL_RUN = r"[^\w\s.,()+\-]{1,3}"
_CURRENCY = re.compile(rf"^({_SYMBOL_RUN})?\s*(.*?)\s*({_SYMBOL_RUN})?$", re.S)
_DIGITS = re.compile(r"^\d+$")
_MINUS_SIGNS = {"−": "-", "–": "-"}


class _Column:
    """One column: where it starts, and the edge its cells share.

    A column anchored on the right is a numeric column; one anchored on the left
    is a text column. Both anchors can be present. A column with neither is one
    a vertical rule proved, which no row occupies by construction.
    """

    __slots__ = ("x", "left_anchor", "right_anchor")

    def __init__(self, x: float, left_anchor=None, right_anchor=None):
        self.x = x
        self.left_anchor = left_anchor
        self.right_anchor = right_anchor

    def occupied_by(self, segment) -> bool:
        if self.left_anchor is not None and abs(segment.rect[0] - self.left_anchor) <= COLUMN_TOLERANCE:
            return True
        return (
            self.right_anchor is not None
            and abs(segment.rect[2] - self.right_anchor) <= COLUMN_TOLERANCE
        )


class _Line:
    """One clustered line of drawn text."""

    __slots__ = ("baseline", "segments")

    def __init__(self, baseline: float, segments: list):
        self.baseline = baseline
        self.segments = segments

    @property
    def size(self) -> float:
        return max((s.size for s in self.segments), default=0.0)

    @property
    def right(self) -> float:
        return max(s.rect[2] for s in self.segments)

    @property
    def text(self) -> str:
        return " ".join(s.text for s in sorted(self.segments, key=lambda s: s.rect[0]))


class _Region:
    """One detected table: its rows, its columns, and what evidence found it."""

    __slots__ = ("page", "lines", "columns", "evidence", "caption", "merged")

    def __init__(self, page, lines, columns, evidence, caption):
        self.page = page
        self.lines = lines
        self.columns = columns
        self.evidence = evidence
        self.caption = caption
        self.merged = 0


# --------------------------------------------------------------------------
# Channels
# --------------------------------------------------------------------------


def _classify_rules(shapes, page_box) -> tuple[list, list]:
    """The page's strokes as (horizontal rules, vertical rules).

    A closed rectangle is decomposed into its four edges rather than kept whole:
    a detector consumes rule POSITIONS, and a cell drawn as one rectangle
    operator carries the same two column boundaries as the same cell drawn as
    four lines.
    """
    page_area = max((page_box[2] - page_box[0]) * (page_box[3] - page_box[1]), 1.0)
    horizontal: list[tuple] = []
    vertical: list[tuple] = []
    for shape in shapes:
        x0, y0, x1, y1 = shape.rect
        width, height = x1 - x0, y1 - y0
        if width <= 0 or height <= 0:
            continue
        if width * height >= page_area * BORDER_AREA_FRACTION:
            continue  # the page border, not a table edge
        thin = max(shape.line_width * 1.5, RULE_THICKNESS_FLOOR)
        if height <= thin and width >= MIN_RULE_WIDTH:
            horizontal.append(shape.rect)
            continue
        if width <= thin and height >= MIN_RULE_WIDTH:
            vertical.append(shape.rect)
            continue
        if shape.kind == "fill" or not shape.closed or shape.points != 4:
            continue
        if width >= MIN_RULE_WIDTH:
            horizontal.append((x0, y0, x1, y0))
            horizontal.append((x0, y1, x1, y1))
        if height >= MIN_RULE_WIDTH:
            vertical.append((x0, y0, x0, y1))
            vertical.append((x1, y0, x1, y1))
    return horizontal, vertical


def _is_vertical_writing(segment) -> bool:
    """A run drawn down the page rather than across it.

    Its glyphs stack, so its left edge says nothing about a column and its width
    says nothing about a span.
    """
    width = segment.rect[2] - segment.rect[0]
    height = segment.rect[3] - segment.rect[1]
    return len(segment.text) > 1 and height > 1.5 * segment.size and width <= 1.5 * segment.size


def _cluster_lines(segments) -> list[_Line]:
    ordered = sorted(segments, key=lambda s: (-s.rect[1], s.rect[0]))
    lines: list[_Line] = []
    for segment in ordered:
        tolerance = max(ROW_TOLERANCE_FLOOR, ROW_TOLERANCE_EM * segment.size)
        if lines and abs(lines[-1].baseline - segment.rect[1]) <= tolerance:
            lines[-1].segments.append(segment)
            continue
        lines.append(_Line(segment.rect[1], [segment]))
    return lines


# --------------------------------------------------------------------------
# Columns and regions
# --------------------------------------------------------------------------


def _edge_clusters(lines, index: int) -> list[tuple[float, int]]:
    """(edge position, how many LINES share it) for one edge of every segment.

    A line votes once per cluster, so a row whose cell wraps into two segments
    cannot vote twice for its own column.
    """
    edges = sorted(segment.rect[index] for line in lines for segment in line.segments)
    if not edges:
        return []
    clusters: list[list[float]] = [[edges[0]]]
    for edge in edges[1:]:
        if edge - clusters[-1][-1] <= COLUMN_TOLERANCE:
            clusters[-1].append(edge)
        else:
            clusters.append([edge])
    out: list[tuple[float, int]] = []
    for cluster in clusters:
        low, high = cluster[0] - COLUMN_TOLERANCE, cluster[-1] + COLUMN_TOLERANCE
        support = sum(
            1
            for line in lines
            if any(low <= segment.rect[index] <= high for segment in line.segments)
        )
        out.append((cluster[0] if index == 0 else cluster[-1], support))
    return out


def _column_positions(lines, minimum_rows: int) -> list[_Column]:
    """The columns these lines establish, left to right.

    Left anchors are proposed first because a left-aligned column's start IS its
    boundary; a right-anchored column's boundary is derived from the widest cell
    that shares the anchor, which is the leftmost point any of its cells reach.
    """
    proposals: list[_Column] = []
    for anchor, support in _edge_clusters(lines, 0):
        if support >= minimum_rows:
            proposals.append(_Column(anchor, left_anchor=anchor))
    for anchor, support in _edge_clusters(lines, 2):
        if support < minimum_rows:
            continue
        members = [
            segment
            for line in lines
            for segment in line.segments
            if abs(segment.rect[2] - anchor) <= COLUMN_TOLERANCE
        ]
        if not members:
            continue
        proposals.append(
            _Column(min(s.rect[0] for s in members), right_anchor=anchor)
        )

    accepted: list[_Column] = []
    for column in sorted(proposals, key=lambda c: (c.x, c.left_anchor is None)):
        if accepted and column.x - accepted[-1].x < MIN_COLUMN_SEPARATION:
            # The same column found through both of its edges: keep the earlier
            # proposal and let it answer for both alignments.
            previous = accepted[-1]
            if previous.left_anchor is None:
                previous.left_anchor = column.left_anchor
            if previous.right_anchor is None:
                previous.right_anchor = column.right_anchor
            continue
        accepted.append(column)
    return accepted


def _occupied(line, columns) -> set:
    return {
        index
        for index, column in enumerate(columns)
        for segment in line.segments
        if column.occupied_by(segment)
    }


def _place(segment, columns) -> int:
    """The column a segment sits in, once the region is confirmed.

    Placement is total and nearest-boundary: every drawn segment inside a region
    belongs somewhere, and a cell that occupies no anchor still has a position.
    """
    reachable = [i for i, c in enumerate(columns) if c.x <= segment.rect[2] + COLUMN_TOLERANCE]
    pool = reachable or list(range(len(columns)))
    return min(pool, key=lambda i: abs(segment.rect[0] - columns[i].x))


def _split_by_rules(columns, verticals, lines) -> list[_Column]:
    """Columns a vertical rule proves the alignment merged.

    A rule inside the region's span with no column beside it bounds a column
    every row leaves empty; a rule at the region's outer edge bounds nothing.
    """
    if not columns or not verticals:
        return columns
    top = max(line.baseline for line in lines)
    bottom = min(line.baseline for line in lines)
    span_right = max(line.right for line in lines)
    out = list(columns)
    for rect in verticals:
        x = (rect[0] + rect[2]) / 2.0
        if rect[3] < bottom or rect[1] > top:
            continue  # does not stand beside these rows
        if x <= columns[0].x - MIN_COLUMN_SEPARATION or x >= span_right:
            continue  # the region's own outer edge
        if any(abs(existing.x - x) < MIN_COLUMN_SEPARATION for existing in out):
            continue  # corroborates a column already found
        out.append(_Column(x + 1.0))
    return sorted(out, key=lambda c: c.x)


def _row_pitch(band) -> float:
    pitches = [band[i].baseline - band[i + 1].baseline for i in range(len(band) - 1)]
    return min(pitches) if pitches else 0.0


def _absorb_spanning_rows(body, band, columns) -> list:
    """Pull in a header row above the band that one cell spans.

    A cell covering two or more columns occupies exactly one of them, so the
    occupancy test cannot see its row at all; the span is what identifies it.
    Without this a merged header is not part of its own table.
    """
    pitch = _row_pitch(band)
    if pitch <= 0 or len(columns) < 2:
        return band
    index = body.index(band[0])
    while index > 0:
        line = body[index - 1]
        if line.baseline - band[0].baseline > 1.5 * pitch:
            break
        if len(line.segments) != 1:
            break
        segment = line.segments[0]
        if segment.rect[0] < columns[0].x - COLUMN_TOLERANCE:
            break
        start = _place(segment, columns)
        if start + 1 >= len(columns):
            break
        if segment.rect[2] <= columns[start + 1].x + MERGE_OVERLAP:
            break
        band = [line] + band
        index -= 1
    return band


def _caption_for(lines, index, region_lines, columns) -> str | None:
    """The line above a region that names it.

    A caption is set no smaller than the body it captions, which is what
    separates a table's title from a note that happens to sit above it.
    """
    if index == 0:
        return None
    pitch = _row_pitch(region_lines)
    if pitch <= 0:
        return None
    body = max(line.size for line in region_lines)
    top = region_lines[0].baseline
    best = None
    for line in reversed(lines[:index]):
        if line.baseline - top > CAPTION_PITCHES * pitch:
            break
        # Strictly larger: a line set at the body's own size cannot be told
        # from a sentence that happens to sit above the table.
        if len(line.segments) != 1 or line.size <= body:
            continue
        if _occupied(line, columns) - {0}:
            continue  # sits in the body columns: a row, not a title
        if best is None or line.size > best.size:
            best = line
    return best.text.strip() if best is not None else None


def _rules_across(rects, rows, vertical: bool = False) -> bool:
    if not rects:
        return False
    top = max(line.baseline for line in rows)
    bottom = min(line.baseline for line in rows)
    for rect in rects:
        if vertical:
            if rect[1] <= top and rect[3] >= bottom:
                return True
        elif bottom - MIN_COLUMN_SEPARATION <= rect[1] <= top + MIN_COLUMN_SEPARATION:
            return True
    return False


def _regions_on_page(page_number, lines, verticals, horizontals) -> tuple[list, list]:
    """This page's table regions, and the lines no region claimed."""
    body = [line for line in lines if line.segments]
    candidates = _column_positions(body, MIN_COLUMN_ROWS)
    if len(candidates) < MIN_REGION_COLUMNS:
        return [], body

    bands: list[list[_Line]] = []
    current: list[_Line] = []
    for line in body:
        if len(_occupied(line, candidates)) < MIN_REGION_COLUMNS:
            if current:
                bands.append(current)
                current = []
            continue
        # Two tables can sit one under the other with nothing between them; the
        # break is the vertical gap, because a run of rows at one pitch is what
        # a table IS.
        pitch = _row_pitch(current)
        if current and pitch > 0 and current[-1].baseline - line.baseline > 2.0 * pitch:
            bands.append(current)
            current = []
        current.append(line)
    if current:
        bands.append(current)

    regions: list[_Region] = []
    claimed: set[int] = set()
    for band in bands:
        if len(band) < MIN_REGION_ROWS:
            continue
        columns = _column_positions(band, MIN_COLUMN_ROWS)
        if len(columns) < MIN_REGION_COLUMNS:
            continue
        columns = _split_by_rules(columns, verticals, band)
        band = _absorb_spanning_rows(body, band, columns)
        evidence = "aligned"
        if _rules_across(horizontals, band):
            evidence = "ruled" if _rules_across(verticals, band, vertical=True) else "banded"
        index = body.index(band[0])
        region = _Region(page_number, band, columns, evidence, None)
        region.caption = _caption_for(body, index, band, columns)
        regions.append(region)
        for line in band:
            claimed.add(id(line))
        if region.caption is not None:
            for line in body[:index]:
                if line.text.strip() == region.caption:
                    claimed.add(id(line))
    untabled = [line for line in body if id(line) not in claimed]
    return regions, untabled


# --------------------------------------------------------------------------
# Cell text and typing
# --------------------------------------------------------------------------


def logical_text(text: str) -> tuple[str, bool]:
    """(a cell's text in logical order, whether the reordering was proven).

    A segment is assembled left to right by geometry, so page order is VISUAL
    order and a right-to-left cell arrives reversed. The inverse is a CANDIDATE
    — bidi is an involution only for two-level text — so it is proven by
    re-running the forward reordering and requiring the permutation to compose
    to the identity. An unprovable cell keeps the text as drawn and is counted.
    """
    if not bidi.has_strong_rtl(text):
        return text, True
    back = bidi.reconstruct_logical(text, 1)
    if len(back) != len(text):
        return text, False
    logical = "".join(text[i] for i in back)
    _level, forward = bidi.visual_order(logical, 1)
    if len(forward) != len(text) or any(back[forward[v]] != v for v in range(len(text))):
        return text, False
    return logical, True


def _decorations(text: str) -> tuple[str, str, bool]:
    """(the bare number, its currency symbol, whether it is a percentage)."""
    body = text.strip()
    percent = body.endswith("%")
    if percent:
        body = body[:-1].strip()
    symbol = ""
    match = _CURRENCY.match(body)
    if match is not None and match.group(2):
        symbol = (match.group(1) or match.group(3) or "").strip()
        body = match.group(2).strip()
    return body, symbol, percent


def _split_groups(body: str, group: str, point: str):
    """(integer digits, fraction digits) under one separator convention, or None."""
    if body.count(point) > 1:
        return None
    whole, _, fraction = body.partition(point)
    if fraction and not _DIGITS.match(fraction):
        return None
    parts = whole.split(group)
    if not all(_DIGITS.match(part) for part in parts):
        return None
    if len(parts) > 1 and (len(parts[0]) > 3 or any(len(part) != 3 for part in parts[1:])):
        return None
    return "".join(parts), fraction


def numeric_convention(texts) -> str:
    """Which separator the DOCUMENT uses for a decimal point.

    Decided from the document's own cells, never from the interface language: a
    number written `1.200,50` is a number whatever language the reader speaks.
    Only decisive tokens vote — a bare integer, and a single separator followed
    by exactly three digits, parse identically under both conventions and say
    nothing.
    """
    dot = comma = 0
    for text in texts:
        body, _symbol, _percent = _decorations(_normalize_signs(text))
        body = body.strip("()").lstrip("+-")
        if not body:
            continue
        has_dot, has_comma = "." in body, "," in body
        if has_dot and has_comma:
            if body.rfind(".") > body.rfind(","):
                dot += 1
            else:
                comma += 1
        elif has_dot:
            if body.count(".") > 1:
                comma += 1
            elif len(body.rsplit(".", 1)[1]) != 3:
                dot += 1
        elif has_comma:
            if body.count(",") > 1:
                dot += 1
            elif len(body.rsplit(",", 1)[1]) != 3:
                comma += 1
    return "comma" if comma > dot else "dot"


def _normalize_signs(text: str) -> str:
    for sign, plain in _MINUS_SIGNS.items():
        text = text.replace(sign, plain)
    return text


def parse_cell(text: str, convention: str):
    """(value, number format) for a cell that is a number or a date, else None.

    A cell that does not parse cleanly stays a string. A spreadsheet whose
    figures are all text is a spreadsheet nobody can total, and a spreadsheet
    whose strings were guessed into numbers is worse.
    """
    stripped = text.strip()
    if not stripped:
        return None
    iso = _ISO_DATE.match(stripped)
    if iso is not None:
        try:
            return (
                datetime.date(int(iso.group(1)), int(iso.group(2)), int(iso.group(3))),
                "yyyy-mm-dd",
            )
        except ValueError:
            return None

    body, symbol, percent = _decorations(_normalize_signs(stripped))
    negative = False
    if body.startswith("(") and body.endswith(")"):
        negative, body = True, body[1:-1].strip()
    if body.startswith("-"):
        negative, body = True, body[1:].strip()
    elif body.startswith("+"):
        body = body[1:].strip()
    if not body:
        return None

    group, point = (".", ",") if convention == "comma" else (",", ".")
    split = _split_groups(body, group, point)
    if split is None:
        return None
    whole, fraction = split
    if not whole:
        return None
    grouped = group in body
    value = float(f"{whole}.{fraction}") if fraction else float(whole)
    if negative:
        value = -value
    decimals = len(fraction)

    if percent:
        return value / 100.0, "0.00%" if decimals else "0%"
    number = ("#,##0" if grouped else "0") + ("." + "0" * decimals if decimals else "")
    if symbol:
        return value, '"{}"{}'.format(symbol.replace('"', ""), number)
    if not fraction and not grouped:
        return int(value), "General"
    return value, number


# --------------------------------------------------------------------------
# The grid
# --------------------------------------------------------------------------


def _grid(region) -> tuple[list[list[str]], list[tuple], int]:
    """(cell text by row and column, merged spans, unresolved right-to-left cells).

    A merged cell is written as one span, never duplicated into each column it
    covers and never truncated to the first.
    """
    rows: list[list[str]] = []
    merges: list[tuple] = []
    unresolved = 0
    for row_index, line in enumerate(region.lines):
        cells = [""] * len(region.columns)
        for segment in sorted(line.segments, key=lambda s: s.rect[0]):
            index = _place(segment, region.columns)
            text, proven = logical_text(segment.text.strip())
            if not proven:
                unresolved += 1
            cells[index] = f"{cells[index]} {text}".strip() if cells[index] else text
            width = 1
            for further in range(index + 1, len(region.columns)):
                if segment.rect[2] <= region.columns[further].x + MERGE_OVERLAP:
                    break
                width += 1
            if width > 1:
                merges.append((row_index, index, width))
        rows.append(cells)
    return rows, merges, unresolved


# --------------------------------------------------------------------------
# Detection and export
# --------------------------------------------------------------------------


def _page_numbers(pages, pdf) -> list[int]:
    """The 1-based pages to analyze.

    Spelled with `page_no` and `len(pdf.pages)` deliberately: the out-of-range
    refusal is a shared row in the engine-message table, and a differently named
    local would rename the interpolations of a message several modules raise.
    """
    if pages is None or pages == "all":
        return list(range(1, len(pdf.pages) + 1))
    if isinstance(pages, str):
        raise ValueError('pages must be a list of page numbers or "all"')
    out: list[int] = []
    for value in pages:
        page_no = int(value)
        if not (1 <= page_no <= len(pdf.pages)):
            raise ValueError(f"page {page_no} is out of range (1-{len(pdf.pages)})")
        if page_no not in out:
            out.append(page_no)
    return sorted(out)


def detect_tables(file: str, pages="all") -> dict:
    """Every table region in scope, plus the lines no region claimed."""
    regions: list[_Region] = []
    untabled: dict[int, list[str]] = {}
    vertical_writing = 0
    with pikepdf.open(str(file)) as pdf:
        wanted = _page_numbers(pages, pdf)
        for number in wanted:
            page = pdf.pages[number - 1]
            segments = _page_segments(pdf, page)
            upright = [s for s in segments if not _is_vertical_writing(s)]
            vertical_writing += len(segments) - len(upright)
            shapes, _placed = _page_shapes(pdf, page)
            horizontals, verticals = _classify_rules(shapes, _crop_box(page))
            lines = _cluster_lines(upright)
            found, spare = _regions_on_page(number, lines, verticals, horizontals)
            regions.extend(found)
            spare_text = [line.text.strip() for line in spare if line.text.strip()]
            if spare_text:
                untabled[number] = spare_text
    return {
        "pages": wanted,
        "regions": regions,
        "untabled": untabled,
        "vertical_writing_runs": vertical_writing,
    }


def _sheet_name(wanted: str, taken: set) -> str:
    base = _SHEET_NAME_BANNED.sub(" ", wanted).strip()[:SHEET_NAME_LIMIT].strip() or "Table"
    name = base
    suffix = 2
    while name.casefold() in taken:
        tail = f"_{suffix}"
        name = base[: SHEET_NAME_LIMIT - len(tail)] + tail
        suffix += 1
    taken.add(name.casefold())
    return name


def _write_region(sheet, region, convention, start_row: int) -> tuple[int, int]:
    """Lay one region out from `start_row`; return (next free row, unresolved)."""
    rows, merges, unresolved = _grid(region)
    for row_index, cells in enumerate(rows):
        for column_index, text in enumerate(cells):
            if not text:
                continue
            cell = sheet.cell(row=start_row + row_index, column=column_index + 1)
            typed = parse_cell(text, convention)
            if typed is None:
                cell.value = text
            else:
                cell.value, cell.number_format = typed
    for row_index, column_index, width in merges:
        region.merged += 1
        sheet.merge_cells(
            start_row=start_row + row_index,
            start_column=column_index + 1,
            end_row=start_row + row_index,
            end_column=column_index + width,
        )
    return start_row + len(rows), unresolved


def _report(region, sheet_title: str, start_row: int) -> dict:
    return {
        "page": region.page,
        "sheet": sheet_title,
        "first_row": start_row,
        "rows": len(region.lines),
        "columns": len(region.columns),
        "evidence": region.evidence,
        "merged": region.merged,
    }


def export_tables(
    file: str,
    output: str,
    pages="all",
    sheet_per: str = "table",
    include_untabled: bool = False,
) -> dict:
    """Write ``file``'s detected tables to the workbook at ``output``.

    Args:
        file: input PDF path.
        output: destination ``.xlsx`` path.
        pages: list of 1-based page numbers, or 'all'.
        sheet_per: 'table' (one sheet per region) or 'page'.
        include_untabled: append a sheet carrying the text no region claimed.
    """
    from openpyxl import Workbook

    grouping = str(sheet_per or "table").lower()
    if grouping not in SHEET_PER:
        raise ValueError(f"unknown sheet grouping {sheet_per!r} (choose table or page)")

    found = detect_tables(file, pages)
    regions = found["regions"]
    wanted = found["pages"]
    if not regions:
        raise ValueError(
            f"no table was found on the {len(wanted)} page(s) analyzed, so there is "
            "nothing to write to a spreadsheet"
        )

    convention = numeric_convention(
        [segment.text for region in regions for line in region.lines for segment in line.segments]
    )

    book = Workbook()
    book.remove(book.active)
    taken: set = set()
    reported: list[dict] = []
    counters: dict[int, int] = {}
    unresolved = 0

    if grouping == "page":
        for number in wanted:
            on_page = [r for r in regions if r.page == number]
            if not on_page:
                continue
            sheet = book.create_sheet(_sheet_name(f"Page_{number}", taken))
            row = 1
            for region in on_page:
                start = row
                row, count = _write_region(sheet, region, convention, row)
                row += 1
                unresolved += count
                reported.append(_report(region, sheet.title, start))
    else:
        for region in regions:
            counters[region.page] = counters.get(region.page, 0) + 1
            stem = region.caption or f"Table_p{region.page}_{counters[region.page]}"
            sheet = book.create_sheet(_sheet_name(stem, taken))
            _row, count = _write_region(sheet, region, convention, 1)
            unresolved += count
            reported.append(_report(region, sheet.title, 1))

    untabled_lines = sum(len(v) for v in found["untabled"].values())
    if include_untabled and untabled_lines:
        sheet = book.create_sheet(_sheet_name("Other text", taken))
        row = 1
        for number in wanted:
            for text in found["untabled"].get(number, []):
                sheet.cell(row=row, column=1, value=number)
                sheet.cell(row=row, column=2, value=logical_text(text)[0])
                row += 1

    out_path = Path(output)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    book.save(str(out_path))

    return {
        "output": str(out_path),
        "format": "xlsx",
        "size": out_path.stat().st_size,
        "pages_analyzed": wanted,
        "tables": reported,
        "untabled_lines": untabled_lines,
        "pages_without_tables": [n for n in wanted if not any(r.page == n for r in regions)],
        "vertical_writing_runs": found["vertical_writing_runs"],
        "unresolved_rtl_cells": unresolved,
        "number_convention": convention,
    }
