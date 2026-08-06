"""Table detection and the workbook it writes."""

import os

import pikepdf
import pytest
from openpyxl import load_workbook
from pikepdf import Array, Dictionary, Name

from engine.office_export import export_document
from engine.table_export import detect_tables, numeric_convention, parse_cell

ROWS = [
    ["Region", "Q1", "Q2", "Q3"],
    ["North", "1200", "1310", "1455"],
    ["South", "980", "1024", "1190"],
    ["East", "1500", "1490", "1610"],
    ["West", "745", "820", "905"],
]
COL_X = [72.0, 250.0, 350.0, 450.0, 540.0]
ROW_Y = [700.0, 676.0, 652.0, 628.0, 604.0, 580.0]
HEBREW = "אבגדהוזחטיכלמנסעפצקרשת"


def _esc(text):
    return text.replace("\\", r"\\").replace("(", r"\(").replace(")", r"\)")


def _write(path, pages_ops, size=(612, 792), rotate=0, hebrew=False):
    pdf = pikepdf.new()
    for ops in pages_ops:
        fonts = {
            "F1": pdf.make_indirect(Dictionary(
                Type=Name.Font, Subtype=Name.Type1,
                BaseFont=Name.Helvetica, Encoding=Name.WinAnsiEncoding))
        }
        if hebrew:
            fonts["F2"] = _hebrew_font(pdf)
        page = Dictionary(
            Type=Name.Page, MediaBox=Array([0, 0, size[0], size[1]]),
            Resources=Dictionary(Font=Dictionary(**fonts)),
            Contents=pdf.make_stream("\n".join(ops).encode("latin-1")))
        if rotate:
            page[Name.Rotate] = rotate
        pdf.pages.append(pikepdf.Page(pdf.make_indirect(page)))
    pdf.save(str(path))
    pdf.close()
    return str(path)


def _hebrew_font(pdf):
    """Helvetica with a /ToUnicode that spells Hebrew.

    The drawn glyphs are Latin and the extracted characters are Hebrew, which is
    the channel detection reads: what a viewer shows is irrelevant to what the
    content stream says the text IS.
    """
    pairs = "".join(f"<{ord('a') + i:02x}> <{ord(ch):04x}>\n" for i, ch in enumerate(HEBREW))
    cmap = (
        "/CIDInit /ProcSet findresource begin 12 dict begin begincmap\n"
        "/CMapName /Custom def /CMapType 2 def\n"
        "1 begincodespacerange <00> <ff> endcodespacerange\n"
        f"{len(HEBREW)} beginbfchar\n{pairs}endbfchar\n"
        "endcmap CMapName currentdict /CMap defineresource pop end end"
    ).encode("latin-1")
    return pdf.make_indirect(Dictionary(
        Type=Name.Font, Subtype=Name.Type1, BaseFont=Name.Helvetica,
        Encoding=Name.WinAnsiEncoding, ToUnicode=pdf.make_stream(cmap)))


def _drawn_rtl(logical):
    """The byte string a viewer draws for `logical`, i.e. its VISUAL order."""
    return "".join(chr(ord("a") + HEBREW.index(ch)) for ch in reversed(logical))


def _cell_ops(rows=ROWS, col_x=COL_X, row_y=ROW_Y):
    ops = []
    for r, row in enumerate(rows):
        base = row_y[r + 1] + 8.0
        for c, cell in enumerate(row):
            ops.append(f"BT /F1 10 Tf {col_x[c] + 4:.1f} {base:.1f} Td ({_esc(cell)}) Tj ET")
    return ops


def _rule_ops(vertical=True):
    ops = ["0.8 w 0 G"]
    for y in ROW_Y:
        ops.append(f"{COL_X[0]} {y} m {COL_X[-1]} {y} l S")
    if vertical:
        for x in COL_X:
            ops.append(f"{x} {ROW_Y[-1]} m {x} {ROW_Y[0]} l S")
    return ops


def _unruled_ops():
    """One show operator per row; the columns are drawn spaces and nothing else."""
    ops = []
    for r, row in enumerate(ROWS):
        base = ROW_Y[r + 1] + 8.0
        text = ""
        for c, cell in enumerate(row):
            pad = max(1, int((COL_X[c] - COL_X[0]) / 5.0) - len(text))
            text += " " * pad + cell
        ops.append(f"BT /F1 10 Tf 72 {base:.1f} Td ({_esc(text)}) Tj ET")
    return ops


def _read_back(path):
    book = load_workbook(path)
    sheet = book.worksheets[0]
    return [[sheet.cell(row=r, column=c).value for c in range(1, 5)] for r in range(1, 6)]


EXPECTED = [
    ["Region", "Q1", "Q2", "Q3"],
    ["North", 1200, 1310, 1455],
    ["South", 980, 1024, 1190],
    ["East", 1500, 1490, 1610],
    ["West", 745, 820, 905],
]


@pytest.mark.parametrize(
    "name,ops",
    [
        ("ruled", _rule_ops(True) + _cell_ops()),
        ("banded", _rule_ops(False) + _cell_ops()),
        ("unruled", _unruled_ops()),
    ],
)
def test_the_same_twenty_cells_come_back_at_the_same_addresses(tmp_dir, name, ops):
    src = _write(os.path.join(tmp_dir, f"{name}.pdf"), [ops])
    out = os.path.join(tmp_dir, f"{name}.xlsx")
    result = export_document(src, out, "xlsx")
    assert len(result["tables"]) == 1
    assert (result["tables"][0]["rows"], result["tables"][0]["columns"]) == (5, 4)
    assert _read_back(out) == EXPECTED


def test_the_unruled_case_depends_on_the_segment_split(tmp_dir, monkeypatch):
    # The mutation test for the unruled fixture: one show operator per row means
    # the only thing separating the columns is the wide-gap segment split. With
    # it disabled the row collapses into one cell and detection must fail.
    import engine.form_detect as fd

    src = _write(os.path.join(tmp_dir, "unruled.pdf"), [_unruled_ops()])
    monkeypatch.setattr(fd, "SEGMENT_GAP_SPACES", 10_000.0)
    with pytest.raises(ValueError, match="no table was found"):
        export_document(src, os.path.join(tmp_dir, "o.xlsx"), "xlsx")


def test_right_aligned_numbers_form_columns(tmp_dir):
    # A numeric column shares its RIGHT edge and no two of its cells start at
    # the same x, so left-edge clustering alone finds nothing there.
    right = [240.0, 340.0, 440.0]
    ops = []
    for r, row in enumerate(ROWS):
        base = ROW_Y[r + 1] + 8.0
        ops.append(f"BT /F1 10 Tf 76 {base:.1f} Td ({_esc(row[0])}) Tj ET")
        for c, cell in enumerate(row[1:]):
            ops.append(
                f"BT /F1 10 Tf {right[c] - len(cell) * 5.0:.1f} {base:.1f} "
                f"Td ({_esc(cell)}) Tj ET")
    src = _write(os.path.join(tmp_dir, "right.pdf"), [ops])
    out = os.path.join(tmp_dir, "right.xlsx")
    export_document(src, out, "xlsx")
    assert _read_back(out) == EXPECTED


def test_a_rotated_page_yields_the_cells_of_its_unrotated_twin(tmp_dir):
    ops = _rule_ops(True) + _cell_ops()
    flat = _write(os.path.join(tmp_dir, "flat.pdf"), [ops])
    turned = _write(os.path.join(tmp_dir, "turned.pdf"), [ops], rotate=90)
    a = os.path.join(tmp_dir, "flat.xlsx")
    b = os.path.join(tmp_dir, "turned.xlsx")
    export_document(flat, a, "xlsx")
    export_document(turned, b, "xlsx")
    assert _read_back(a) == _read_back(b) == EXPECTED


def test_a_merged_header_is_one_span_not_two_cells(tmp_dir):
    ops = ["BT /F1 10 Tf 76 692 Td (Region and quarter summary spanning columns) Tj ET"]
    ops += _cell_ops(rows=ROWS[1:], row_y=ROW_Y[1:])
    src = _write(os.path.join(tmp_dir, "merged.pdf"), [ops])
    out = os.path.join(tmp_dir, "merged.xlsx")
    result = export_document(src, out, "xlsx")
    assert result["tables"][0]["merged"] == 1
    book = load_workbook(out)
    sheet = book.worksheets[0]
    assert [str(r) for r in sheet.merged_cells.ranges] == ["A1:B1"]
    assert sheet["A1"].value == "Region and quarter summary spanning columns"
    assert sheet["B1"].value is None
    assert sheet["A2"].value == "North"


def test_a_vertical_rule_splits_a_column_the_alignment_merged(tmp_dir):
    col_x = [72.0, 200.0, 320.0, 440.0, 540.0]
    ops = ["0.8 w 0 G"]
    for y in ROW_Y:
        ops.append(f"{col_x[0]} {y} m {col_x[-1]} {y} l S")
    for x in col_x:
        ops.append(f"{x} {ROW_Y[-1]} m {x} {ROW_Y[0]} l S")
    for r, row in enumerate(ROWS):
        base = ROW_Y[r + 1] + 8.0
        for c, cell in enumerate([row[0], row[1], "", row[3]]):
            if cell:
                ops.append(
                    f"BT /F1 10 Tf {col_x[c] + 4:.1f} {base:.1f} Td ({_esc(cell)}) Tj ET")
    src = _write(os.path.join(tmp_dir, "gap.pdf"), [ops])
    out = os.path.join(tmp_dir, "gap.xlsx")
    result = export_document(src, out, "xlsx")
    assert result["tables"][0]["columns"] == 4
    book = load_workbook(out)
    sheet = book.worksheets[0]
    assert [sheet.cell(row=1, column=c).value for c in range(1, 5)] == [
        "Region", "Q1", None, "Q3"
    ]


def test_two_regions_on_one_page_become_two_sheets(tmp_dir):
    ops = _cell_ops()
    ops.append("BT /F1 10 Tf 72 552 Td (Prose between the two tables goes here.) Tj ET")
    second_x = [72.0, 260.0, 420.0]
    second_y = [520.0, 496.0, 472.0, 448.0, 424.0]
    second = [["Item", "Team", "State"], ["Alpha", "Dana", "open"],
              ["Beta", "Ravi", "closed"], ["Gamma", "Iris", "open"]]
    ops += _cell_ops(rows=second, col_x=second_x, row_y=second_y)
    src = _write(os.path.join(tmp_dir, "two.pdf"), [ops])
    out = os.path.join(tmp_dir, "two.xlsx")
    result = export_document(src, out, "xlsx")
    shapes = [(t["rows"], t["columns"]) for t in result["tables"]]
    assert shapes == [(5, 4), (4, 3)]
    assert len(load_workbook(out).worksheets) == 2
    assert result["untabled_lines"] == 1


def test_untabled_text_is_counted_and_can_be_carried(tmp_dir):
    ops = ["BT /F1 10 Tf 72 730 Td (All figures in thousands.) Tj ET"]
    ops += _rule_ops(True) + _cell_ops()
    ops.append("BT /F1 10 Tf 72 560 Td (Totals exclude intercompany transfers.) Tj ET")
    src = _write(os.path.join(tmp_dir, "notes.pdf"), [ops])
    counted = os.path.join(tmp_dir, "counted.xlsx")
    carried = os.path.join(tmp_dir, "carried.xlsx")
    plain = export_document(src, counted, "xlsx")
    assert plain["untabled_lines"] == 2
    assert len(load_workbook(counted).worksheets) == 1
    export_document(src, carried, "xlsx", include_untabled=True)
    book = load_workbook(carried)
    assert len(book.worksheets) == 2
    spare = book.worksheets[-1]
    assert [spare.cell(row=r, column=2).value for r in (1, 2)] == [
        "All figures in thousands.",
        "Totals exclude intercompany transfers.",
    ]


def test_pages_without_tables_are_named_not_refused(tmp_dir):
    src = _write(
        os.path.join(tmp_dir, "mixed.pdf"),
        [
            _rule_ops(True) + _cell_ops(),
            ["BT /F1 11 Tf 72 700 Td (A page of ordinary prose text.) Tj ET"],
        ],
    )
    result = export_document(src, os.path.join(tmp_dir, "mixed.xlsx"), "xlsx")
    assert result["pages_analyzed"] == [1, 2]
    assert result["pages_without_tables"] == [2]
    assert len(result["tables"]) == 1


def test_sheet_per_page_lays_every_region_down_one_sheet(tmp_dir):
    ops = _cell_ops()
    second_y = [520.0, 496.0, 472.0, 448.0, 424.0]
    second = [["Item", "Team", "State"], ["Alpha", "Dana", "open"],
              ["Beta", "Ravi", "closed"], ["Gamma", "Iris", "open"]]
    ops += _cell_ops(rows=second, col_x=[72.0, 260.0, 420.0], row_y=second_y)
    src = _write(os.path.join(tmp_dir, "grouped.pdf"), [ops])
    out = os.path.join(tmp_dir, "grouped.xlsx")
    result = export_document(src, out, "xlsx", sheet_per="page")
    book = load_workbook(out)
    assert len(book.worksheets) == 1
    assert [t["first_row"] for t in result["tables"]] == [1, 7]
    sheet = book.worksheets[0]
    assert sheet["A1"].value == "Region"
    assert sheet["A7"].value == "Item"


def test_rejects_an_unknown_sheet_grouping(tmp_dir):
    src = _write(os.path.join(tmp_dir, "s.pdf"), [_cell_ops()])
    with pytest.raises(ValueError, match="unknown sheet grouping"):
        export_document(src, os.path.join(tmp_dir, "o.xlsx"), "xlsx", sheet_per="workbook")


def test_a_document_with_no_table_refuses_by_name(tmp_dir):
    src = _write(
        os.path.join(tmp_dir, "prose.pdf"),
        [["BT /F1 11 Tf 72 700 Td (Nothing here resembles a table at all.) Tj ET"]],
    )
    out = os.path.join(tmp_dir, "prose.xlsx")
    with pytest.raises(ValueError, match="no table was found on the 1 page"):
        export_document(src, out, "xlsx")
    # A refusal is a RESULT: no workbook survives for a caller to mistake for one.
    assert not os.path.exists(out)


def test_a_page_of_photographs_has_no_tables(tmp_dir):
    pdf = pikepdf.new()
    image = pdf.make_stream(bytes([200, 30, 30] * 64))
    image.Type = Name.XObject
    image.Subtype = Name.Image
    image.Width = 8
    image.Height = 8
    image.ColorSpace = Name.DeviceRGB
    image.BitsPerComponent = 8
    pdf.pages.append(pikepdf.Page(pdf.make_indirect(Dictionary(
        Type=Name.Page, MediaBox=Array([0, 0, 612, 792]),
        Resources=Dictionary(XObject=Dictionary(Im0=image)),
        Contents=pdf.make_stream(b"q 400 0 0 400 100 200 cm /Im0 Do Q")))))
    src = os.path.join(tmp_dir, "photo.pdf")
    pdf.save(src)
    pdf.close()
    with pytest.raises(ValueError, match="no table was found"):
        export_document(src, os.path.join(tmp_dir, "o.xlsx"), "xlsx")


def test_a_right_to_left_table_reads_left_to_right_with_logical_cells(tmp_dir):
    rows = [["אבג", "דהו", "זחט"], ["יכל", "מנס", "עפצ"],
            ["קרש", "תאב", "גדה"], ["וזח", "טיכ", "למנ"]]
    xs = [72.0, 200.0, 330.0]
    ys = [700.0, 676.0, 652.0, 628.0]
    ops = []
    for r, row in enumerate(rows):
        for c, cell in enumerate(row):
            ops.append(
                f"BT /F2 10 Tf {xs[c]:.1f} {ys[r] - 8:.1f} Td ({_drawn_rtl(cell)}) Tj ET")
    src = _write(os.path.join(tmp_dir, "rtl.pdf"), [ops], hebrew=True)
    out = os.path.join(tmp_dir, "rtl.xlsx")
    result = export_document(src, out, "xlsx")
    assert result["unresolved_rtl_cells"] == 0
    book = load_workbook(out)
    sheet = book.worksheets[0]
    # Column A is the leftmost column the page drew, and each cell reads in
    # logical order rather than as the page laid its characters out.
    assert [sheet.cell(row=1, column=c).value for c in (1, 2, 3)] == ["אבג", "דהו", "זחט"]
    assert sheet["A4"].value == "וזח"


def test_vertical_writing_is_excluded_from_columns_and_reported(tmp_dir):
    ops = _cell_ops()
    for n in range(3):
        ops.append(f"BT /F1 10 Tf 0 1 -1 0 {560 + n * 14} 600 Tm (Sidebar note) Tj ET")
    src = _write(os.path.join(tmp_dir, "vertical.pdf"), [ops])
    out = os.path.join(tmp_dir, "vertical.xlsx")
    result = export_document(src, out, "xlsx")
    assert result["vertical_writing_runs"] == 3
    assert _read_back(out) == EXPECTED


def test_sheet_names_come_from_the_caption_and_deduplicate(tmp_dir):
    ops = ["BT /F1 14 Tf 72 730 Td (Quarterly revenue) Tj ET"]
    ops += _cell_ops()
    second_y = [520.0, 496.0, 472.0, 448.0, 424.0]
    ops.append("BT /F1 14 Tf 72 546 Td (Quarterly revenue) Tj ET")
    ops += _cell_ops(rows=ROWS[:4], col_x=COL_X, row_y=second_y)
    src = _write(os.path.join(tmp_dir, "named.pdf"), [ops])
    out = os.path.join(tmp_dir, "named.xlsx")
    export_document(src, out, "xlsx")
    titles = [s.title for s in load_workbook(out).worksheets]
    assert titles == ["Quarterly revenue", "Quarterly revenue_2"]


def test_a_long_caption_is_truncated_to_the_formats_limit(tmp_dir):
    caption = "A caption far longer than a sheet name may ever be"
    ops = [f"BT /F1 14 Tf 72 730 Td ({_esc(caption)}) Tj ET"] + _cell_ops()
    src = _write(os.path.join(tmp_dir, "long.pdf"), [ops])
    out = os.path.join(tmp_dir, "long.xlsx")
    export_document(src, out, "xlsx")
    title = load_workbook(out).worksheets[0].title
    assert len(title) <= 31
    assert caption.startswith(title)


def test_detect_tables_refuses_a_page_outside_the_document(tmp_dir):
    src = _write(os.path.join(tmp_dir, "s.pdf"), [_cell_ops()])
    with pytest.raises(ValueError, match=r"page 4 is out of range \(1-1\)"):
        detect_tables(src, [4])


# -- numeric typing ---------------------------------------------------------


@pytest.mark.parametrize(
    "text,value,fmt",
    [
        ("1200", 1200, "General"),
        ("1,200", 1200.0, "#,##0"),
        ("$1,200.50", 1200.5, '"$"#,##0.00'),
        ("-42.5", -42.5, "0.0"),
        ("(1,200)", -1200.0, "#,##0"),
        ("12.5%", 0.125, "0.00%"),
        ("50%", 0.5, "0%"),
    ],
)
def test_a_number_is_written_as_a_number(text, value, fmt):
    assert parse_cell(text, "dot") == (value, fmt)


@pytest.mark.parametrize("text", ["Q1", "N/A", "", "12/31/2026", "1,20", "abc", "1.2.3"])
def test_anything_that_does_not_parse_cleanly_stays_a_string(text):
    assert parse_cell(text, "dot") is None


def test_a_date_is_written_as_a_date():
    import datetime

    assert parse_cell("2019-03-07", "dot") == (datetime.date(2019, 3, 7), "yyyy-mm-dd")
    assert parse_cell("2019-13-07", "dot") is None


def test_the_separator_convention_comes_from_the_document():
    assert numeric_convention(["1,200.50", "980"]) == "dot"
    assert numeric_convention(["1.200,50", "980"]) == "comma"
    # A bare integer and a lone three-digit group are identical under both
    # conventions, so neither votes and the default stands.
    assert numeric_convention(["980", "1,200"]) == "dot"
    assert parse_cell("1.200,50", "comma") == (1200.5, "#,##0.00")
    assert parse_cell("12,5%", "comma") == (0.125, "0.00%")
